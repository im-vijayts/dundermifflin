import openpyxl
import random

class GetDialogue():
    def __init__(self):
        self.wb = openpyxl.load_workbook('michael.xlsx')
        sheets = self.wb.sheetnames
        self.ws = self.wb[sheets[0]]
        self.no_of_row = self.ws.max_row + 1
        self.dialogue_col_no = 5
    def getDialogue(self):
        rowno = random.randint(2, self.no_of_row)
        data = self.ws.cell(rowno, self.dialogue_col_no).value
        return data